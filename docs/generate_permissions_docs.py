#!/usr/bin/env python3
"""Generate Excel + PDF for Aretia Permissions & Roles (client share)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from fpdf import FPDF

DOCS = Path(__file__).resolve().parent
XLSX_OUT = DOCS / "Aretia-Permissions-and-Roles.xlsx"
PDF_OUT = DOCS / "Aretia-Permissions-and-Roles.pdf"

ROLES = ["Client", "Analyst", "QA", "FQA", "Admin", "Super Admin"]
YES = "Yes"
NO = "No"

HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="EEF2FF")
SECTION_FONT = Font(bold=True, size=11, color="312E81")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def yn(val: str) -> str:
    if val in ("✓", "Yes", YES):
        return YES
    if val in ("—", "No", NO, "-"):
        return NO
    return val


MATRIX = [
    ("Portal & account", [
        ("Log in to portal", "✓", "✓", "✓", "✓", "✓", "✓", ""),
        ("Edit own profile", "✓", "✓", "✓", "✓", "✓", "✓", ""),
        ("Notifications & chat inbox", "✓", "✓", "✓", "✓", "✓", "✓", ""),
        ("Permissions & Roles settings", "—", "—", "—", "—", "—", "✓", "Super Admin only; UI coming soon"),
    ]),
    ("Onboarding & companies", [
        ("Complete onboarding (KYC)", "✓", "—", "—", "—", "—", "—", "Client only"),
        ("Review / approve / reject onboarding", "—", "—", "—", "—", "✓", "✓", ""),
        ("View clients / suspend company", "—", "—", "—", "—", "✓", "✓", ""),
    ]),
    ("Orders", [
        ("View orders", "✓", "—", "—", "—", "✓", "✓", "Client: whole company"),
        ("Create / import orders", "✓", "—", "—", "—", "✓", "✓", "Client: whole company"),
        ("Upload order documents", "✓", "—", "—", "—", "✓", "✓", ""),
        ("Approve / reject order", "—", "—", "—", "—", "✓", "✓", ""),
        ("Set order due date", "✓", "—", "—", "—", "✓", "✓", ""),
    ]),
    ("Cases", [
        ("View / open cases", "✓", "✓", "✓", "✓", "✓", "✓", "Client: company; Staff: assigned or all"),
        ("Assign case team", "—", "—", "—", "—", "✓", "✓", ""),
        ("Link related cases", "✓", "—", "—", "—", "✓", "✓", "Client: company"),
        ("Upload / download case documents", "✓", "✓", "✓", "✓", "✓", "✓", ""),
    ]),
    ("Case workflow", [
        ("Advance stage (workflow lanes)", "—", "✓", "✓", "✓", "—", "—", "See Workflow sheet"),
        ("Override any stage", "—", "—", "—", "—", "✓", "✓", ""),
    ]),
    ("Reports", [
        ("Deliver report to client", "—", "—", "—", "✓", "✓", "✓", "FQA when stage = FQA started"),
        ("View / download reports", "✓", "—", "—", "—", "—", "—", "Client: company"),
    ]),
    ("Messages & chat", [
        ("Case chat (after analyst assigned)", "✓", "✓", "✓", "✓", "✓", "✓", ""),
        ("Notifications to all company clients", "✓", "—", "—", "—", "—", "—", "When staff sends message"),
        ("Notifications to case team", "—", "✓", "✓", "✓", "✓", "✓", "When client sends message"),
    ]),
    ("People & access", [
        ("Manage employees", "—", "—", "—", "—", "✓", "✓", ""),
        ("Manage client users", "—", "—", "—", "—", "✓", "✓", ""),
        ("Manage Super Admin accounts", "—", "—", "—", "—", "—", "—", "Protected"),
    ]),
    ("Workflow & audit", [
        ("Configure workflow stages", "—", "—", "—", "—", "✓", "✓", ""),
        ("Operations dashboard", "✓", "✓", "✓", "✓", "✓", "✓", ""),
        ("Audit trail", "—", "—", "—", "—", "✓", "✓", ""),
    ]),
]

WORKFLOW = [
    ("Analyst", "Assigned -> Research started -> Research done"),
    ("QA", "Research done -> QA started -> QA done"),
    ("FQA", "QA done -> FQA started -> Sent to client"),
]

ROLE_OVERVIEW = [
    ("Client", "Company customer", "Orders, cases, reports, chat for their company"),
    ("Analyst", "Research lead", "Assigned cases; research workflow stages"),
    ("QA", "Quality review", "Assigned cases; QA stages after research done"),
    ("FQA", "Final QA", "Assigned cases; delivers report to client"),
    ("Admin", "Operations", "Onboarding, orders, cases, teams, full platform ops"),
    ("Super Admin", "Platform owner", "Same as Admin + future system settings"),
]


def build_workbook() -> Workbook:
    wb = Workbook()

    # --- Roles ---
    ws = wb.active
    ws.title = "Roles"
    ws.append(["Role", "Label", "Purpose"])
    for r in ROLE_OVERVIEW:
        ws.append(list(r))
    style_header_row(ws, 1)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 55

    # --- Permission matrix ---
    ws2 = wb.create_sheet("Permissions")
    headers = ["Section", "Permission"] + ROLES + ["Notes"]
    ws2.append(headers)
    style_header_row(ws2, 1)
    row = 2
    for section, rows in MATRIX:
        ws2.cell(row, 1, section)
        ws2.cell(row, 1).font = SECTION_FONT
        ws2.cell(row, 1).fill = SECTION_FILL
        for col in range(1, len(headers) + 1):
            ws2.cell(row, col).border = BORDER
        row += 1
        for perm_row in rows:
            cells = [section, perm_row[0]] + [yn(perm_row[i]) for i in range(1, 7)] + [perm_row[7]]
            ws2.append(cells)
            for c in range(3, 9):
                ws2.cell(row, c).alignment = CENTER
            row += 1
        row += 1

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 38
    for col in "CDEFGH":
        ws2.column_dimensions[col].width = 12
    ws2.column_dimensions["I"].width = 36
    ws2.freeze_panes = "C2"

    # --- Workflow ---
    ws3 = wb.create_sheet("Workflow stages")
    ws3.append(["Role", "Allowed stage progression"])
    style_header_row(ws3, 1)
    for r in WORKFLOW:
        ws3.append(list(r))
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 55

    # --- Legend ---
    ws4 = wb.create_sheet("Legend")
    ws4.append(["Symbol", "Meaning"])
    style_header_row(ws4, 1)
    legend = [
        ("Yes", "Permission granted for that role"),
        ("No", "Permission not granted"),
        ("Company scope", "All client users under the same company name see shared orders/cases"),
        ("Assigned", "Employee must be on the case team"),
        ("Lane", "Employee may only move stages in their workflow lane"),
    ]
    for item in legend:
        ws4.append(list(item))
    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 60

    return wb


def style_header_row(ws, row_num: int) -> None:
    for cell in ws[row_num]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER


def build_pdf() -> None:
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Aretia - Permissions & Roles", ln=True)
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(
        0,
        5,
        "Who can do what in the portal. Yes = allowed, No = not allowed. "
        "Client users in the same company share orders, cases, chat, and reports.",
    )
    pdf.ln(3)

    col_w = [42, 78] + [22] * 6 + [50]
    total = sum(col_w)
    scale = (277 - 14) / total
    col_w = [w * scale for w in col_w]

    def table_header():
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(79, 70, 229)
        pdf.set_text_color(255, 255, 255)
        headers = ["Section", "Permission"] + ROLES + ["Notes"]
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 7, h[:18], border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

    table_header()
    pdf.set_font("Helvetica", "", 6.5)
    fill_toggle = False
    for section, rows in MATRIX:
        for perm_row in rows:
            if pdf.get_y() > 185:
                pdf.add_page()
                table_header()
                pdf.set_font("Helvetica", "", 6.5)
            fill = fill_toggle
            fill_toggle = not fill_toggle
            if fill:
                pdf.set_fill_color(248, 250, 252)
            data = [section[:20], perm_row[0][:42]] + [yn(perm_row[i]) for i in range(1, 7)] + [
                (perm_row[7] or "")[:28]
            ]
            for i, txt in enumerate(data):
                pdf.cell(col_w[i], 5, txt, border=1, fill=fill, align="C" if i >= 2 and i <= 7 else "L")
            pdf.ln()

    pdf.add_page()
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, "Workflow stage lanes (employees)", ln=True)
    pdf.set_font("Helvetica", "", 10)
    for role, path in WORKFLOW:
        pdf.cell(0, 6, f"  {role}: {path}", ln=True)

    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(pdf.epw, 8, "Role summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 9)
    for role, label, purpose in ROLE_OVERVIEW:
        pdf.multi_cell(pdf.epw, 5, f"{role} ({label}): {purpose}")
        pdf.ln(1)

    pdf.output(str(PDF_OUT))


def main() -> None:
    wb = build_workbook()
    wb.save(XLSX_OUT)
    build_pdf()
    print(f"Created: {XLSX_OUT}")
    print(f"Created: {PDF_OUT}")


if __name__ == "__main__":
    main()
